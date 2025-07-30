"""
File processing functions for TRAK file generation.
"""
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from price_calculator import get_cd_price, get_lp_price

def process_input_data(input_file_path, output_dir):
    """
    Process the input data and create a complete Excel file.
    
    Args:
        input_file_path (str): Path to the input Excel file
        output_dir (str): Directory to save the output file
        
    Returns:
        tuple: (DataFrame with processed data, path to the Excel output file)
    """
    # Define required and optional columns
    required_columns = ['UPC', 'TITLE', 'ARTIST', 'MANUF', 'GENRE', 'CONFIG', 'COST']
    optional_columns = ['DEPT', 'MISC', 'LIST', 'PRICE', 'VENDOR']
    all_columns = required_columns + optional_columns
    
    # Define data types for columns - keep text fields as strings, allow numeric fields as floats
    dtype_dict = {
        'UPC': str,      # Always string to preserve exact format
        'TITLE': str,
        'ARTIST': str,
        'MANUF': str,
        'GENRE': str,
        'CONFIG': str,
        'DEPT': str,     # Always string to preserve exact format
        'MISC': str,
        'VENDOR': str
        # COST, LIST, PRICE are not specified so they can be read as floats naturally
    }
    
    # Read the Excel file with specified data types
    df = pd.read_excel(input_file_path, dtype=dtype_dict)
    
    # Check for missing columns and add them if they don't exist
    for column in all_columns:
        if column not in df.columns:
            if column in dtype_dict:
                df[column] = ''  # String columns get empty string
            else:
                df[column] = None  # Numeric columns get None (will become NaN)

    # Drop rows where all specified columns are NaN (blank)
    df = df.dropna(subset=all_columns, how='all')

    # Replace NaN values with an empty string
    df = df.fillna('')

    # Set MISC and VENDOR to MANUF if they're not provided
    df['MISC'] = df.apply(lambda row: row['MANUF'] if not row['MISC'] else row['MISC'], axis=1)
    df['VENDOR'] = df.apply(lambda row: row['MANUF'] if not row['VENDOR'] else row['VENDOR'], axis=1)

    # Remove dollar signs from PRICE, COST, and LIST columns
    df['PRICE'] = df['PRICE'].replace({r'\$': ''}, regex=True)
    df['COST'] = df['COST'].replace({r'\$': ''}, regex=True)
    df['LIST'] = df['LIST'].replace({r'\$': ''}, regex=True)

    # Process each row to fill in missing values
    processed_rows = []
    
    for index, row in df.iterrows():
        processed_row = row.copy()
        
        # Pad UPC with leading zeros to ensure it's at least 12 digits
        if processed_row['UPC'] and processed_row['UPC'] != 'nan' and str(processed_row['UPC']).strip():
            # UPC is already a string, preserve it as much as possible
            upc = str(processed_row['UPC']).strip()
            
            # Only remove non-digit characters if absolutely necessary
            # First check if it's already all digits
            if upc.isdigit():
                # UPC is already clean, just pad if needed
                pass
            else:
                # Only filter if there are non-digit characters
                upc = ''.join(filter(str.isdigit, upc))
            
            # Pad with leading zeros if less than 12 digits
            if len(upc) < 12:
                upc = upc.zfill(12)
            processed_row['UPC'] = upc
        
        # Set department based on CONFIG if DEPT is empty
        if not processed_row['DEPT'] or processed_row['DEPT'] == 'nan' or str(processed_row['DEPT']).strip() == '':
            if processed_row['CONFIG'] == 'CD':
                processed_row['DEPT'] = '02'
            elif processed_row['CONFIG'] == 'LP':
                processed_row['DEPT'] = '01'
        else:
            # DEPT is already a string, ensure it's formatted as two digits
            dept_str = str(processed_row['DEPT']).strip()
            try:
                dept_num = int(dept_str)
                # Always format as two digits with leading zero (01, 02, etc.)
                processed_row['DEPT'] = f"{dept_num:02d}"
            except (ValueError, TypeError):
                # If DEPT contains non-numeric values, keep as is
                processed_row['DEPT'] = dept_str
                
        # Calculate LIST and PRICE based on CONFIG if missing
        try:
            cost_value = float(processed_row['COST']) if processed_row['COST'] else 0
        except (ValueError, TypeError):
            cost_value = 0
        
        if processed_row['CONFIG'] == 'CD':
            # If price is missing, determine it from the cost
            if not processed_row['PRICE']:
                calculated_price = get_cd_price(cost_value)
                if calculated_price:
                    processed_row['PRICE'] = format(calculated_price, '.2f')
            
            # If list price is missing, use the same calculation as price
            if not processed_row['LIST']:
                calculated_list = get_cd_price(cost_value)
                if calculated_list:
                    processed_row['LIST'] = format(calculated_list, '.2f')
                    
        elif processed_row['CONFIG'] == 'LP':
            # If price is missing, determine it from the cost using the LP pricing table
            if not processed_row['PRICE']:
                calculated_price = get_lp_price(cost_value)
                if calculated_price:
                    processed_row['PRICE'] = format(calculated_price, '.2f')
            
            # If list price is missing, use the same calculation as price
            if not processed_row['LIST']:
                calculated_list = get_lp_price(cost_value)
                if calculated_list:
                    processed_row['LIST'] = format(calculated_list, '.2f')
        
        # Format cost value
        if processed_row['COST']:
            try:
                processed_row['COST'] = format(float(processed_row['COST']), '.2f')
            except (ValueError, TypeError):
                # If COST contains non-numeric values, keep as is
                pass
            
        processed_rows.append(processed_row)
    
    # Create a new DataFrame with processed data
    processed_df = pd.DataFrame(processed_rows)
    
    # Reorder columns to match the desired output format
    ordered_columns = ['UPC', 'TITLE', 'ARTIST', 'MANUF', 'GENRE', 'CONFIG', 'DEPT', 'MISC', 'LIST', 'PRICE', 'VENDOR', 'COST']
    for col in ordered_columns:
        if col not in processed_df.columns:
            processed_df[col] = ''  # Ensure all columns exist
    processed_df = processed_df[ordered_columns]
    
    # Create output file paths
    excel_output_path = os.path.join(output_dir, 'trakdelim.xlsx')
    
    # Save to a new Excel file with proper formatting
    with pd.ExcelWriter(excel_output_path, engine='openpyxl') as writer:
        processed_df.to_excel(writer, index=False, sheet_name='Sheet1')
        
        # Get the workbook and worksheet to format columns
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        # Format UPC and DEPT columns as text to preserve leading zeros and exact format
        from openpyxl.styles import NamedStyle
        text_style = NamedStyle(name="text_style")
        text_style.number_format = '@'  # Text format
        
        # Apply text formatting to UPC column (column A)
        for row in range(2, len(processed_df) + 2):  # Start from row 2 (after header)
            worksheet[f'A{row}'].style = text_style
            
        # Apply text formatting to DEPT column (column G)
        for row in range(2, len(processed_df) + 2):  # Start from row 2 (after header)
            worksheet[f'G{row}'].style = text_style
    
    return processed_df, excel_output_path

def generate_delimited_file(df, output_dir):
    """
    Generate the delimited text file from the processed data.
    
    Args:
        df (DataFrame): Processed data (can be passed directly or read from Excel)
        output_dir (str): Directory to save the output file
        
    Returns:
        str: Path to the generated text file
    """
    # If df is a file path (string), read it with proper data types
    if isinstance(df, str):
        # Define data types to preserve text formatting
        dtype_dict = {
            'UPC': str,
            'TITLE': str,
            'ARTIST': str,
            'MANUF': str,
            'GENRE': str,
            'CONFIG': str,
            'DEPT': str,     # Keep as string to preserve 01, 02 format
            'MISC': str,
            'VENDOR': str
        }
        df = pd.read_excel(df, dtype=dtype_dict)
    
    # Create output file path
    text_output_path = os.path.join(output_dir, 'trakdelim.txt')
    
    # Open the output file for writing in the specified directory
    with open(text_output_path, 'w') as file:
        
        # Iterate over each row in the DataFrame
        for index, row in df.iterrows():
            # Process monetary values to ensure they always have cents positions without the decimal point
            
            # Process LIST column - format as whole number with implied decimal (e.g., 1100 for $11.00)
            if pd.notna(row['LIST']) and row['LIST'] != '' and str(row['LIST']).lower() != 'nan':
                try:
                    # Convert to float, format with 2 decimal places, then remove the decimal point
                    list_value = float(str(row['LIST']).replace('$', '').strip())
                    list_price = f"{list_value:.2f}".replace('.', '')
                except (ValueError, TypeError):
                    list_price = ''
            else:
                list_price = ''
                
            # Process PRICE column - format as whole number with implied decimal (e.g., 950 for $9.50)
            if pd.notna(row['PRICE']) and row['PRICE'] != '' and str(row['PRICE']).lower() != 'nan':
                try:
                    # Convert to float, format with 2 decimal places, then remove the decimal point
                    price_value = float(str(row['PRICE']).replace('$', '').strip())
                    price = f"{price_value:.2f}".replace('.', '')
                except (ValueError, TypeError):
                    price = ''
            else:
                price = ''
                
            # Process COST column - format as whole number with implied decimal (e.g., 1000 for $10.00)
            if pd.notna(row['COST']) and row['COST'] != '' and str(row['COST']).lower() != 'nan':
                try:
                    # Convert to float, format with 2 decimal places, then remove the decimal point
                    cost_value = float(str(row['COST']).replace('$', '').strip())
                    cost = f"{cost_value:.2f}".replace('.', '')
                except (ValueError, TypeError):
                    cost = ''
            else:
                cost = ''
            
            # Get values for each field, using empty string for any that are missing or 'nan'
            # Ensure UPC is preserved exactly as it is in the spreadsheet, including leading zeros
            upc = str(row.get('UPC', ''))
            upc = '' if upc.lower() == 'nan' else upc
            
            title = str(row.get('TITLE', ''))
            title = '' if title.lower() == 'nan' else title
            
            artist = str(row.get('ARTIST', ''))
            artist = '' if artist.lower() == 'nan' else artist
            
            manuf = str(row.get('MANUF', ''))
            manuf = '' if manuf.lower() == 'nan' else manuf
            
            genre = str(row.get('GENRE', ''))
            genre = '' if genre.lower() == 'nan' else genre
            
            misc = str(row.get('MISC', ''))
            misc = '' if misc.lower() == 'nan' else misc
            
            config = str(row.get('CONFIG', ''))
            config = '' if config.lower() == 'nan' else config
            
            dept = str(row.get('DEPT', ''))
            dept = '' if dept.lower() == 'nan' else dept
            
            vendor = str(row.get('VENDOR', ''))
            vendor = '' if vendor.lower() == 'nan' else vendor
            
            # Format the row data according to the specified layout
            formatted_row = (
                f"C|{upc}|{title}|{artist}|{manuf}|||{genre}|||{misc}|{config}|||{dept}|{list_price}||||||{vendor}|{cost}|||||||||{price}"
            )

            # Write the formatted data to the output file
            file.write(formatted_row + '\n')
            
    return text_output_path

def create_new_spreadsheet(file_path):
    """
    Create a new spreadsheet with required and optional columns.
    
    Args:
        file_path (str): Path where the new spreadsheet should be saved
        
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        # Create a new DataFrame with the required columns
        required_columns = ['UPC', 'TITLE', 'ARTIST', 'MANUF', 'GENRE', 'CONFIG', 'COST']
        optional_columns = ['DEPT', 'MISC', 'LIST', 'PRICE', 'VENDOR']
        all_columns = required_columns + optional_columns
        
        # Create an empty DataFrame with the columns
        df = pd.DataFrame(columns=all_columns)
        
        # Save to Excel
        df.to_excel(file_path, index=False)
        
        # Format the Excel file with styles
        workbook = load_workbook(file_path)
        sheet = workbook.active
        
        # Format the header row - required columns in red, optional in black
        for col_idx, column_name in enumerate(all_columns, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            if column_name in required_columns:
                cell.font = Font(color="FF0000", bold=True)  # Red
            else:
                cell.font = Font(bold=True)  # Black, bold
        
        # Save the formatted workbook
        workbook.save(file_path)
        return True
    except Exception:
        return False